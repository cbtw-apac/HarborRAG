from __future__ import annotations

from io import BytesIO
from typing import Any, ClassVar

from harborrag_core.domain.element import DocumentElement
from harborrag_core.domain.parser import ParsedDocument, ParseInput

from .base import BaseParser
from .exceptions import ParseError
from .parser_logging import get_parser_logger, input_label, parser_log_extra
from .utils import guard_input_size, wrap_parse_errors

parser_logger = get_parser_logger("excel")


class ExcelParser(BaseParser[ParseInput, ParsedDocument]):
    """Extract workbook sheets as tab-separated table text."""

    parser_name: ClassVar[str] = "excel"
    parser_engine: ClassVar[str] = "openpyxl/xlrd"
    suffixes: ClassVar[frozenset[str]] = frozenset({"xls", "xlsx", "xlsm", "xltx", "xltm"})
    content_types: ClassVar[frozenset[str]] = frozenset(
        {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "application/vnd.ms-excel.sheet.macroenabled.12",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.template",
            "application/vnd.ms-excel.template.macroenabled.12",
        }
    )

    def parse(self, input: ParseInput) -> ParsedDocument:
        """Route legacy `.xls` through xlrd and OpenXML workbooks through openpyxl."""

        parse_input = self.coerce_input(input)
        source_bytes = guard_input_size(parse_input.read_bytes())
        parser_logger.debug(
            "Starting Excel parse for %s",
            input_label(parse_input),
            extra=parser_log_extra(
                input=parse_input,
                parser_name=self.parser_name,
                parser_engine=self.parser_engine,
                input_bytes=len(source_bytes),
            ),
        )

        if parse_input.suffix == ".xls":
            content, elements, sheet_names = self._parse_xls(parse_input, source_bytes)
        else:
            content, elements, sheet_names = self._parse_openxml(
                parse_input,
                source_bytes,
            )

        parser_logger.info(
            "Parsed Excel workbook %s sheets=%d content_chars=%d elements=%d",
            input_label(parse_input),
            len(sheet_names),
            len(content),
            len(elements),
            extra=parser_log_extra(
                input=parse_input,
                parser_name=self.parser_name,
                parser_engine=self.parser_engine,
                sheets=len(sheet_names),
                content_chars=len(content),
                elements=len(elements),
            ),
        )
        return ParsedDocument(
            content=content,
            elements=elements,
            parser_name=self.parser_name,
            parser_version=self.parser_version,
            metadata=self.metadata_for(parse_input, sheets=sheet_names),
        )

    def _parse_openxml(
        self,
        parse_input: ParseInput,
        source_bytes: bytes,
    ) -> tuple[str, list[DocumentElement], list[str]]:
        """Read `.xlsx`-style workbooks in read-only, data-only mode."""

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            parser_logger.error(
                "Excel parser dependency `openpyxl` is missing",
                extra=parser_log_extra(
                    input=parse_input,
                    parser_name=self.parser_name,
                    parser_engine=self.parser_engine,
                ),
            )
            raise ParseError(
                "Excel parsing requires `openpyxl`; install `harborrag-adapters[parsers]` "
                "or `pip install openpyxl`."
            ) from exc

        parser_logger.debug(
            "Extracting OpenXML Excel text from %s",
            input_label(parse_input),
            extra=parser_log_extra(
                input=parse_input,
                parser_name=self.parser_name,
                parser_engine="openpyxl",
                input_bytes=len(source_bytes),
            ),
        )
        with wrap_parse_errors("openpyxl"):
            workbook = load_workbook(
                BytesIO(source_bytes),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        sheet_names = workbook.sheetnames
        try:
            sections: list[str] = []
            elements: list[DocumentElement] = []
            for sheet in workbook.worksheets:
                rows = [
                    "\t".join(self._cell_to_text(value) for value in row).rstrip()
                    for row in sheet.iter_rows(values_only=True)
                ]
                rows = [row for row in rows if row.strip()]
                if not rows:
                    parser_logger.debug(
                        "Skipping empty Excel sheet %s",
                        sheet.title,
                        extra=parser_log_extra(
                            input=parse_input,
                            parser_name=self.parser_name,
                            parser_engine="openpyxl",
                            sheet=sheet.title,
                            rows=0,
                        ),
                    )
                    continue
                sheet_text = "\n".join(rows)
                sections.append(f"Sheet: {sheet.title}\n{sheet_text}")
                elements.append(
                    DocumentElement(
                        id=f"excel:sheet:{sheet.title}",
                        type="table",
                        content=sheet_text,
                        metadata={"sheet": sheet.title},
                    )
                )
                parser_logger.debug(
                    "Extracted Excel sheet %s rows=%d content_chars=%d",
                    sheet.title,
                    len(rows),
                    len(sheet_text),
                    extra=parser_log_extra(
                        input=parse_input,
                        parser_name=self.parser_name,
                        parser_engine="openpyxl",
                        sheet=sheet.title,
                        rows=len(rows),
                        content_chars=len(sheet_text),
                    ),
                )
        finally:
            workbook.close()

        return "\n\n".join(sections).strip(), elements, sheet_names

    def _parse_xls(
        self,
        parse_input: ParseInput,
        source_bytes: bytes,
    ) -> tuple[str, list[DocumentElement], list[str]]:
        """Read legacy binary `.xls` workbooks with xlrd."""

        try:
            import xlrd
        except ImportError as exc:
            parser_logger.error(
                "Excel parser dependency `xlrd` is missing",
                extra=parser_log_extra(
                    input=parse_input,
                    parser_name=self.parser_name,
                    parser_engine=self.parser_engine,
                ),
            )
            raise ParseError(
                "Legacy .xls parsing requires `xlrd`; install "
                "`harborrag-adapters[parsers]` or `pip install xlrd`."
            ) from exc

        parser_logger.debug(
            "Extracting legacy Excel text from %s",
            input_label(parse_input),
            extra=parser_log_extra(
                input=parse_input,
                parser_name=self.parser_name,
                parser_engine="xlrd",
                input_bytes=len(source_bytes),
            ),
        )
        with wrap_parse_errors("xlrd"):
            workbook = xlrd.open_workbook(
                file_contents=source_bytes,
                on_demand=True,
            )
        sheet_names = workbook.sheet_names()
        sections: list[str] = []
        elements: list[DocumentElement] = []
        try:
            # Load sheets one at a time (Book.sheets() would defeat on_demand and
            # load them all), unloading each after rendering to bound memory.
            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                rows = []
                for row_index in range(sheet.nrows):
                    row = "\t".join(
                        self._xls_cell_to_text(
                            sheet.cell(row_index, column_index),
                            workbook.datemode,
                            xlrd,
                        )
                        for column_index in range(sheet.ncols)
                    ).rstrip()
                    if row.strip():
                        rows.append(row)
                if not rows:
                    parser_logger.debug(
                        "Skipping empty legacy Excel sheet %s",
                        sheet.name,
                        extra=parser_log_extra(
                            input=parse_input,
                            parser_name=self.parser_name,
                            parser_engine="xlrd",
                            sheet=sheet.name,
                            rows=0,
                        ),
                    )
                    workbook.unload_sheet(sheet_index)
                    continue
                sheet_text = "\n".join(rows)
                sections.append(f"Sheet: {sheet.name}\n{sheet_text}")
                elements.append(
                    DocumentElement(
                        id=f"excel:sheet:{sheet.name}",
                        type="table",
                        content=sheet_text,
                        metadata={"sheet": sheet.name},
                    )
                )
                parser_logger.debug(
                    "Extracted legacy Excel sheet %s rows=%d content_chars=%d",
                    sheet.name,
                    len(rows),
                    len(sheet_text),
                    extra=parser_log_extra(
                        input=parse_input,
                        parser_name=self.parser_name,
                        parser_engine="xlrd",
                        sheet=sheet.name,
                        rows=len(rows),
                        content_chars=len(sheet_text),
                    ),
                )
                workbook.unload_sheet(sheet_index)
        finally:
            workbook.release_resources()

        return "\n\n".join(sections).strip(), elements, sheet_names

    @staticmethod
    def _cell_to_text(value: Any) -> str:
        """Convert openpyxl cell values into stable searchable text."""

        if value is None:
            return ""
        isoformat = getattr(value, "isoformat", None)
        if callable(isoformat):
            return isoformat()
        return str(value)

    @staticmethod
    def _xls_cell_to_text(cell: Any, datemode: int, xlrd: Any) -> str:
        """Convert xlrd cell values while preserving dates, booleans, and errors."""

        if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, datemode).isoformat()
            except (OverflowError, ValueError):
                return str(cell.value)
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            number = float(cell.value)
            return str(int(number)) if number.is_integer() else str(number)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if cell.value else "FALSE"
        if cell.ctype == xlrd.XL_CELL_ERROR:
            return f"#ERROR:{cell.value}"
        return str(cell.value)
