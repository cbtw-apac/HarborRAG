"""A worksheet whose <dimension> spans the whole sheet must still parse."""

from __future__ import annotations

import re
import zipfile
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from harborrag_adapters.parsers.common.validation import ParseResourceBudget
from harborrag_adapters.parsers.errors import ParseError
from harborrag_adapters.parsers.spreadsheet.engines.openpyxl import engine as excel_engine
from harborrag_adapters.parsers.spreadsheet.engines.openpyxl.rendering import (
    open_data_workbook,
)
from harborrag_core.domain.parser import ParseInput

_WHOLE_SHEET_ROWS = 'ref="A1:Z1048576"'
_WHOLE_SHEET_COLUMNS = 'ref="A1:XFD2"'


def _workbook_bytes(*, declared_dimension: str | None, extra_rows: int = 0) -> bytes:
    """Build a tiny workbook, optionally lying about its dimension.

    Tool-generated workbooks routinely emit a <dimension> covering the entire sheet, which
    is what a Confluence attachment did in production. Patching the record is how the test
    reproduces that without shipping the customer's file. A lie about the column count is
    the same lie about a different axis, so both are patchable here.
    """

    book = Workbook()
    sheet = book.active
    sheet.title = "data"
    sheet.append(["Story ID", "Objective"])
    sheet.append(["S-1", "Onboard"])
    for index in range(extra_rows):
        sheet.append([f"S-{index + 2}", "Filler"])
    buffer = BytesIO()
    book.save(buffer)
    original = buffer.getvalue()
    if declared_dimension is None:
        return original

    source = zipfile.ZipFile(BytesIO(original))
    patched = BytesIO()
    with zipfile.ZipFile(patched, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename.startswith("xl/worksheets/"):
                payload = re.sub(
                    rb'ref="[A-Z0-9:]+"', declared_dimension.encode(), payload, count=1
                )
            out.writestr(item, payload)
    return patched.getvalue()


def test_a_whole_sheet_row_dimension_resolves_to_the_real_shape() -> None:
    """Read-only mode pads iter_rows out to the declaration, so trusting it both rejects
    the file as hostile and walks a million empty rows. Measured in production: declared
    1048576 x 26, really 790 x 26, parser_rejected_document on every ingestion."""

    data = _workbook_bytes(declared_dimension=_WHOLE_SHEET_ROWS)

    lying = load_workbook(BytesIO(data), read_only=True, data_only=True)
    assert lying.worksheets[0].max_row == 1048576
    lying.close()

    workbook = open_data_workbook(load_workbook, data)
    try:
        sheet = workbook.worksheets[0]
        # The declaration is dropped rather than recomputed, which is what keeps the
        # workbook lazy: a non-streaming reload would materialize every declared cell
        # before ParseResourceBudget ever sees a row.
        assert workbook.read_only is True
        assert sheet.max_row is None
        rows = [row for row in sheet.iter_rows(values_only=True) if any(row)]
        assert rows[0] == ("Story ID", "Objective")
        assert len(rows) == 2
    finally:
        workbook.close()


def test_a_whole_sheet_column_dimension_resolves_to_the_real_shape() -> None:
    """The same lie on the column axis: XFD is column 16384, and read-only mode pads every
    row out to it, so an honest two-column sheet reads as a 16384-cell row."""

    data = _workbook_bytes(declared_dimension=_WHOLE_SHEET_COLUMNS)

    lying = load_workbook(BytesIO(data), read_only=True, data_only=True)
    assert lying.worksheets[0].max_column == 16384
    lying.close()

    workbook = open_data_workbook(load_workbook, data)
    try:
        sheet = workbook.worksheets[0]
        assert workbook.read_only is True
        assert sheet.max_column is None
        rows = [row for row in sheet.iter_rows(values_only=True) if any(row)]
        assert rows == [("Story ID", "Objective"), ("S-1", "Onboard")]
    finally:
        workbook.close()


def test_an_honest_dimension_keeps_the_declared_shape() -> None:
    """The repair is only for the uninformative case; everything else is left alone."""

    workbook = open_data_workbook(load_workbook, _workbook_bytes(declared_dimension=None))
    try:
        assert workbook.read_only is True
        assert workbook.worksheets[0].max_row == 2
    finally:
        workbook.close()


@pytest.mark.parametrize("declared_dimension", [None, _WHOLE_SHEET_ROWS, _WHOLE_SHEET_COLUMNS])
def test_the_same_cells_are_read_either_way(declared_dimension: str | None) -> None:
    workbook = open_data_workbook(
        load_workbook, _workbook_bytes(declared_dimension=declared_dimension)
    )
    try:
        rows = [row for row in workbook.worksheets[0].iter_rows(values_only=True) if any(row)]
    finally:
        workbook.close()

    assert rows == [("Story ID", "Objective"), ("S-1", "Onboard")]


def test_a_repaired_sheet_is_still_rejected_when_it_exceeds_the_budget(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Dropping the declaration also drops the up-front dimension guard, so the per-row
    budget has to carry the limit alone -- which is the point of staying lazy: the reject
    lands after a bounded number of rows rather than after the whole sheet is in memory.
    """

    monkeypatch.setattr(
        excel_engine,
        "ParseResourceBudget",
        lambda: ParseResourceBudget(max_rows=4),
    )
    data = _workbook_bytes(declared_dimension=_WHOLE_SHEET_ROWS, extra_rows=20)

    with pytest.raises(ParseError, match="row count"):
        excel_engine.ExcelSpreadsheetEngine().parse(ParseInput(content=data, filename="big.xlsx"))
