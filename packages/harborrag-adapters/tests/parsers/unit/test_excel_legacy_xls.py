"""Unit tests for Excel parser legacy .xls behavior."""

from __future__ import annotations

import io

import pytest

from harborrag_core.domain.parser import ParseInput

pytestmark = [pytest.mark.unit, pytest.mark.graybox]


def _xls_bytes() -> bytes:
    import xlwt

    book = xlwt.Workbook()
    sheet = book.add_sheet("Data")
    for r, row in enumerate([["name", "score"], ["Ada", 42], ["Bob", 3.5]]):
        for c, value in enumerate(row):
            sheet.write(r, c, value)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_excel_parser_reads_legacy_xls() -> None:
    from harborrag_adapters.parsers.spreadsheet.engines.openpyxl.engine import ExcelParser

    doc = ExcelParser().parse(ParseInput(content=_xls_bytes(), filename="legacy.xls"))
    assert "Data" in doc.content
    assert "Ada" in doc.content and "42" in doc.content
    assert doc.metadata["sheets"] == ["Data"]


def test_excel_parser_missing_dependency_raises_parse_error(monkeypatch) -> None:
    import builtins

    from harborrag_adapters.parsers.errors import ParseError
    from harborrag_adapters.parsers.spreadsheet.engines.openpyxl.engine import ExcelParser

    real_import = builtins.__import__

    def fake_import(name, *args, **kwargs):
        if name == "openpyxl":
            raise ImportError("no openpyxl")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", fake_import)
    with pytest.raises(ParseError, match="openpyxl"):
        ExcelParser().parse(ParseInput(content=b"PK\x03\x04", filename="x.xlsx"))


def test_excel_parser_wraps_xlsx_sheet_iteration_failure_as_parse_error(monkeypatch) -> None:
    """`load_workbook(read_only=True)` only reads headers eagerly; a malformed
    sheet fails later, during `iter_rows()`. That call used to sit outside
    `wrap_parse_errors`, so a corrupt-but-not-zip-bomb sheet leaked a raw
    library exception instead of the typed `ParseError` other engines raise."""
    from harborrag_adapters.parsers.errors import ParseError
    from harborrag_adapters.parsers.spreadsheet.engines.openpyxl.engine import ExcelParser

    def _boom(*_args, **_kwargs):
        raise ValueError("corrupt sheet XML")

    monkeypatch.setattr(ExcelParser, "_cell_to_text", staticmethod(_boom))

    xlsx = _openxml_bytes()
    with pytest.raises(ParseError, match="openpyxl failed to parse input"):
        ExcelParser().parse(ParseInput(content=xlsx, filename="book.xlsx"))


def test_excel_parser_wraps_xls_sheet_iteration_failure_as_parse_error(monkeypatch) -> None:
    """`xlrd.open_workbook(on_demand=True)` defers per-sheet parsing to
    `sheet_by_index()`, called during iteration below the `with`, not during
    `open_workbook()` itself -- so a corrupt sheet must still be caught."""
    from harborrag_adapters.parsers.errors import ParseError
    from harborrag_adapters.parsers.spreadsheet.engines.openpyxl.engine import ExcelParser

    def _boom(*_args, **_kwargs):
        raise ValueError("corrupt legacy sheet")

    monkeypatch.setattr(ExcelParser, "_xls_cell_to_text", staticmethod(_boom))

    with pytest.raises(ParseError, match="xlrd failed to parse input"):
        ExcelParser().parse(ParseInput(content=_xls_bytes(), filename="legacy.xls"))


def _openxml_bytes() -> bytes:
    import io

    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.append(["header"])
    workbook.active.append(["value"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
